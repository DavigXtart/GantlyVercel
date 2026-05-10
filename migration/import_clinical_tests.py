#!/usr/bin/env python3
"""
Batch importer for clinical tests (3-sheet Excel format) and TCP test.
Reads Excel files and inserts into PostgreSQL using the TestEntity model.

Usage: python3 import_clinical_tests.py
"""

import os
import re
import json
import sys
import io

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl psycopg2-binary")
    import openpyxl

import psycopg2

# ── DB config ──────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host": "localhost",
    "port": 5433,
    "dbname": "gantly",
    "user": "postgres",
    "password": "1234",
}

# ── Directories ────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
FILES1_DIR = os.path.join(BASE, "test-imports", "files1")
FILES2_DIR = os.path.join(BASE, "test-imports", "files2")
TCP_FILE = os.path.join(BASE, "test-imports", "TCP_test_final_solo.xlsx")


def connect_db():
    return psycopg2.connect(**DB_CONFIG)


def sanitize_code(name, max_len=30):
    """Generate a short code from a subscale name."""
    # Remove accents and special chars
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ñ": "n", "ü": "u", "Á": "A", "É": "E", "Í": "I",
        "Ó": "O", "Ú": "U", "Ñ": "N",
    }
    code = name
    for k, v in replacements.items():
        code = code.replace(k, v)
    # Keep only alphanumeric and spaces
    code = re.sub(r"[^a-zA-Z0-9 ]", "", code)
    # Take first letters of each word if long
    words = code.strip().split()
    if len(code) > max_len:
        code = "".join(w[0:3].upper() for w in words if w)
    else:
        code = code.strip().upper().replace(" ", "_")
    return code[:max_len]


def parse_clinical_excel(filepath):
    """Parse a 3-sheet clinical test Excel file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames

    # ── Sheet 1: INFO GENERAL ──
    info_sheet = wb[sheets[0]]
    info = {}
    for row in info_sheet.iter_rows(min_row=1, max_row=info_sheet.max_row, values_only=True):
        vals = [v for v in row if v is not None]
        if len(vals) >= 2:
            key = str(vals[0]).strip().lower()
            info[key] = str(vals[1]).strip()
        elif len(vals) == 1:
            # First row often has only the title
            if "title" not in info:
                info["title"] = str(vals[0]).strip()

    # Extract metadata
    title = info.get("nombre completo", info.get("title", os.path.basename(filepath)))
    abbreviation = info.get("abreviatura", "")
    category = info.get("area de evaluacion", info.get("categoria", ""))
    description = info.get("descripcion", "")
    instructions = info.get("instrucciones de administracion", "")
    if instructions:
        description = description + "\n\nInstrucciones: " + instructions

    # Clean title emoji prefix
    title = re.sub(r"^[^\w]+", "", title).strip()
    if not title:
        title = abbreviation or os.path.basename(filepath).replace(".xlsx", "")

    # ── Sheet 2: ITEMS Y RESPUESTAS ──
    items_sheet = wb[sheets[1]]
    all_rows = list(items_sheet.iter_rows(min_row=1, max_row=items_sheet.max_row, values_only=True))

    # Detect format: A1 (shared Likert, 7 cols) vs A2 (per-item, 5 cols)
    global_answers = []
    items = []
    items_start_row = 0

    # Check if first row contains "OPCIONES DE RESPUESTA"
    first_cell = str(all_rows[0][0] or "").strip() if all_rows else ""
    is_format_a1 = "OPCIONES DE RESPUESTA" in first_cell.upper() or "opciones de respuesta" in first_cell.lower()

    if is_format_a1:
        # Format A1: shared Likert options at top
        # Find the answer options (rows after "Valor" header until empty row)
        i = 1
        while i < len(all_rows):
            row = all_rows[i]
            val0 = row[0] if row[0] is not None else ""
            val1 = row[1] if len(row) > 1 and row[1] is not None else ""

            if str(val0).strip().lower() in ("valor", "puntuacion"):
                i += 1
                continue

            if val0 == "" or val0 is None:
                # Empty row = end of global answers
                i += 1
                break

            try:
                score = int(val0)
                label = str(val1).strip()
                if label:
                    global_answers.append({"value": score, "text": label})
            except (ValueError, TypeError):
                pass
            i += 1

        # Skip until we find item headers (row with "No" or "Item")
        while i < len(all_rows):
            row = all_rows[i]
            val0 = str(row[0] or "").strip().lower()
            if val0 in ("no", "nº", "n", "item", "no."):
                items_start_row = i + 1
                break
            i += 1
            items_start_row = i
    else:
        # Format A2: headers in first row
        items_start_row = 1

    # Parse items
    for idx in range(items_start_row, len(all_rows)):
        row = all_rows[idx]
        if row is None or all(v is None for v in row):
            continue

        item_no = row[0]
        item_text = row[1] if len(row) > 1 else None
        factor_name = row[2] if len(row) > 2 else None
        scoring = row[3] if len(row) > 3 else None
        specific_options = row[4] if len(row) > 4 else None

        if item_text is None or str(item_text).strip() == "":
            continue
        if item_no is None:
            continue

        # Skip header rows that snuck through
        item_text_str = str(item_text).strip()
        if item_text_str.lower().startswith("item (texto"):
            continue

        try:
            position = int(item_no)
        except (ValueError, TypeError):
            continue

        # Determine if inverse
        is_inverse = False
        if scoring:
            scoring_str = str(scoring).strip().lower()
            is_inverse = "inversa" in scoring_str or "invertida" in scoring_str or "reverse" in scoring_str

        # Determine answers for this item
        item_answers = []
        if specific_options and str(specific_options).strip():
            # Per-item specific options (pipe-separated: "0=Nunca | 1=A veces | ...")
            opts_str = str(specific_options).strip()
            for opt in opts_str.split("|"):
                opt = opt.strip()
                match = re.match(r"(\d+)\s*[=:]\s*(.+)", opt)
                if match:
                    item_answers.append({"value": int(match.group(1)), "text": match.group(2).strip()})
                elif opt:
                    # Try to parse without explicit value
                    item_answers.append({"value": len(item_answers), "text": opt})
        elif global_answers:
            item_answers = list(global_answers)

        # Factor/subscale
        factor = str(factor_name).strip() if factor_name else "General"
        if factor.lower() in ("none", "", "nan"):
            factor = "General"

        items.append({
            "position": position,
            "text": item_text_str,
            "factor": factor,
            "inverse": is_inverse,
            "answers": item_answers,
        })

    # ── Sheet 3: CORRECCION Y PUNTUACION ──
    cutoffs_by_subscale = {}
    if len(sheets) >= 3:
        corr_sheet = wb[sheets[2]]
        corr_rows = list(corr_sheet.iter_rows(min_row=1, max_row=corr_sheet.max_row, values_only=True))

        current_subscale = None
        in_cutoff_table = False

        for row in corr_rows:
            if row is None or all(v is None for v in row):
                in_cutoff_table = False
                continue

            val0 = str(row[0] or "").strip()

            # Detect subscale header patterns:
            # ">> Depresion (items 3,5,...)" or "► Ansiedad" or "SISTEMA DE CORRECCION"
            if val0.startswith(">>") or val0.startswith(">") or val0.startswith("\u25ba"):
                subscale_name = re.sub(r"^[>\u25ba\s]+", "", val0)
                subscale_name = re.sub(r"\(items?.*\)$", "", subscale_name, flags=re.IGNORECASE).strip()
                if subscale_name and not subscale_name.upper().startswith("SISTEMA"):
                    current_subscale = subscale_name
                    cutoffs_by_subscale.setdefault(current_subscale, [])
                    in_cutoff_table = False
                continue

            # Also detect "ESCALA DE..." or "SUBESCALA:" headers
            if re.match(r"^(ESCALA|SUBESCALA|DIMENSION|FACTOR)\s*[:de]", val0, re.IGNORECASE):
                subscale_name = re.sub(r"^(ESCALA|SUBESCALA|DIMENSION|FACTOR)\s*[:de]\s*", "", val0, flags=re.IGNORECASE).strip()
                if subscale_name:
                    current_subscale = subscale_name
                    cutoffs_by_subscale.setdefault(current_subscale, [])
                    in_cutoff_table = False
                continue

            # Detect cutoff table header (various formats)
            val0_lower = val0.lower()
            val1_str = str(row[1] or "").lower() if len(row) > 1 else ""
            if ("puntuacion" in val0_lower or "puntuaci\u00f3n" in val0_lower) and \
               ("desde" in val0_lower or "desde" in val1_str or "hasta" in val1_str or "clasificacion" in val1_str):
                in_cutoff_table = True
                # If no subscale set yet, use "General"
                if current_subscale is None:
                    current_subscale = "General"
                    cutoffs_by_subscale.setdefault("General", [])
                continue

            # Parse cutoff rows (numeric, numeric, text, ...)
            if in_cutoff_table and current_subscale:
                try:
                    min_score = row[0]
                    max_score = row[1]
                    classification = row[2] if len(row) > 2 else None

                    if min_score is not None and max_score is not None and classification:
                        # Try to convert scores to float
                        min_f = float(min_score)
                        max_f = float(max_score)
                        severity = str(row[4]).strip() if len(row) > 4 and row[4] else str(classification).strip()
                        action = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                        cutoffs_by_subscale[current_subscale].append({
                            "min": min_f,
                            "max": max_f,
                            "label": str(classification).strip(),
                            "severity": severity,
                            "action": action,
                        })
                except (ValueError, TypeError):
                    # Not a cutoff row, end of table
                    in_cutoff_table = False

        # If no subscale headers found, try single-subscale approach
        if not cutoffs_by_subscale:
            in_cutoff_table = False
            for row in corr_rows:
                if row is None:
                    continue
                val0 = str(row[0] or "").strip().lower()
                if ("puntuacion" in val0 or "puntuaci\u00f3n" in val0) and \
                   ("desde" in val0 or "hasta" in val0):
                    in_cutoff_table = True
                    cutoffs_by_subscale.setdefault("General", [])
                    continue
                if in_cutoff_table:
                    try:
                        if row[0] is not None and row[1] is not None and row[2]:
                            cutoffs_by_subscale["General"].append({
                                "min": float(row[0]),
                                "max": float(row[1]),
                                "label": str(row[2]).strip(),
                                "severity": str(row[4]).strip() if len(row) > 4 and row[4] else str(row[2]).strip(),
                                "action": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                            })
                    except (ValueError, TypeError):
                        in_cutoff_table = False

    wb.close()

    return {
        "title": title,
        "abbreviation": abbreviation,
        "category": category,
        "description": description,
        "items": items,
        "cutoffs": cutoffs_by_subscale,
    }


def parse_tcp_excel(filepath):
    """Parse the TCP test (single sheet format with factor codes)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))  # skip header

    items = []
    factors_info = {}  # code -> {name, description}

    for i, row in enumerate(rows):
        if row is None or row[0] is None:
            continue

        question_text = str(row[0]).strip()
        if not question_text:
            continue

        # Answers in columns B-F (0-4 points)
        answers = []
        for pts in range(5):
            col_idx = pts + 1  # B=1, C=2, D=3, E=4, F=5
            label = row[col_idx] if len(row) > col_idx and row[col_idx] else f"{pts} puntos"
            answers.append({"value": pts, "text": str(label).strip()})

        factor_code = str(row[6]).strip() if len(row) > 6 and row[6] else "GEN"
        factor_name = str(row[7]).strip() if len(row) > 7 and row[7] else factor_code
        factor_desc = str(row[8]).strip() if len(row) > 8 and row[8] else ""

        if factor_code not in factors_info:
            factors_info[factor_code] = {"name": factor_name, "description": factor_desc}

        items.append({
            "position": i + 1,
            "text": question_text,
            "factor": factor_code,
            "factor_name": factor_name,
            "inverse": False,
            "answers": answers,
        })

    wb.close()

    return {
        "title": "Test de Competencias Personales (TCP)",
        "abbreviation": "TCP",
        "category": "Personalidad",
        "description": "Test de Competencias Personales - 104 items con 20 factores de personalidad",
        "items": items,
        "factors_info": factors_info,
        "cutoffs": {},
    }


def _normalize(s):
    """Normalize string for fuzzy matching."""
    s = s.lower().strip()
    for k, v in {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n","ü":"u"}.items():
        s = s.replace(k, v)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def _match_cutoffs(factor_name, cutoff_data, total_factors):
    """Match cutoffs from CORRECCION sheet to a factor/subfactor name."""
    if not cutoff_data:
        return None

    # 1. Exact match
    matched = cutoff_data.get(factor_name)
    if matched:
        return matched

    # 2. Case-insensitive match
    fn_lower = factor_name.lower().strip()
    for ck, cv in cutoff_data.items():
        if ck.lower().strip() == fn_lower:
            return cv

    # 3. Normalized match (remove accents, special chars)
    fn_norm = _normalize(factor_name)
    for ck, cv in cutoff_data.items():
        if _normalize(ck) == fn_norm:
            return cv

    # 4. Substring match (either direction)
    for ck, cv in cutoff_data.items():
        ck_norm = _normalize(ck)
        if fn_norm in ck_norm or ck_norm in fn_norm:
            return cv

    # 5. Single-factor test: use whatever cutoffs we have
    if total_factors == 1:
        for cv in cutoff_data.values():
            if cv:  # non-empty
                return cv

    # 6. "General" cutoffs as fallback for single-subscale
    if "General" in cutoff_data and total_factors == 1:
        return cutoff_data["General"]

    return None


def insert_test(conn, test_data, code_override=None):
    """Insert a parsed test into the database."""
    cur = conn.cursor()

    code = code_override or test_data["abbreviation"] or sanitize_code(test_data["title"])
    title = test_data["title"]
    category = test_data["category"]
    description = test_data["description"]

    # Check if test already exists
    cur.execute("SELECT id FROM tests WHERE code = %s", (code,))
    existing = cur.fetchone()
    if existing:
        print(f"  SKIP: Test '{code}' already exists (id={existing[0]})")
        cur.close()
        return existing[0]

    # Insert test
    cur.execute(
        "INSERT INTO tests (code, title, description, active, category, topic, created_at) "
        "VALUES (%s, %s, %s, true, %s, %s, NOW()) RETURNING id",
        (code, title[:200], description, category[:100] if category else None, category[:100] if category else None)
    )
    test_id = cur.fetchone()[0]
    print(f"  Created test '{code}' (id={test_id}): {title[:60]}")

    # Collect unique factors/subscales from items
    factor_names = {}
    for item in test_data["items"]:
        fn = item.get("factor", "General")
        if fn not in factor_names:
            factor_names[fn] = len(factor_names)

    # Create subfactors
    subfactor_map = {}  # factor_name -> subfactor_id
    for factor_name, pos in factor_names.items():
        sf_code = sanitize_code(factor_name)
        sf_name = factor_name[:100]

        # Get cutoffs for this subfactor
        cutoffs_json = None
        cutoff_data = test_data.get("cutoffs", {})
        matched_cutoffs = _match_cutoffs(factor_name, cutoff_data, len(factor_names))

        if matched_cutoffs:
            cutoffs_json = json.dumps(matched_cutoffs, ensure_ascii=False)

        cur.execute(
            "INSERT INTO subfactors (test_id, code, name, description, position, cutoffs) "
            "VALUES (%s, %s, %s, %s, %s, %s) RETURNING id",
            (test_id, sf_code[:30], sf_name, None, pos, cutoffs_json)
        )
        sf_id = cur.fetchone()[0]
        subfactor_map[factor_name] = sf_id

    # If TCP format with factors_info, also create a "General" factor grouping
    if "factors_info" in test_data:
        for fc, fi in test_data["factors_info"].items():
            # Subfactors already created via factor_names
            # Update subfactor description
            sf_id = subfactor_map.get(fc)
            if sf_id and fi.get("description"):
                cur.execute(
                    "UPDATE subfactors SET description = %s, name = %s WHERE id = %s",
                    (fi["description"][:500] if fi["description"] else None,
                     fi["name"][:100] if fi["name"] else fc, sf_id)
                )

    # Create questions and answers
    questions_created = 0
    answers_created = 0

    for item in test_data["items"]:
        factor_name = item.get("factor", "General")
        subfactor_id = subfactor_map.get(factor_name)

        cur.execute(
            "INSERT INTO questions (test_id, text, type, position, subfactor_id, inverse) "
            "VALUES (%s, %s, 'SINGLE', %s, %s, %s) RETURNING id",
            (test_id, item["text"][:500], item["position"], subfactor_id, item.get("inverse", False))
        )
        question_id = cur.fetchone()[0]
        questions_created += 1

        for ans_pos, answer in enumerate(item.get("answers", [])):
            cur.execute(
                "INSERT INTO answers (question_id, text, value, position, is_correct) "
                "VALUES (%s, %s, %s, %s, false)",
                (question_id, str(answer["text"])[:500], answer["value"], ans_pos)
            )
            answers_created += 1

    conn.commit()
    print(f"    -> {questions_created} questions, {answers_created} answers, {len(subfactor_map)} subfactors")

    # Count cutoffs stored
    cur.execute("SELECT COUNT(*) FROM subfactors WHERE test_id = %s AND cutoffs IS NOT NULL", (test_id,))
    cutoff_count = cur.fetchone()[0]
    if cutoff_count > 0:
        print(f"    -> {cutoff_count} subfactors with cutoffs")

    cur.close()
    return test_id


def main():
    conn = connect_db()
    print("Connected to PostgreSQL\n")

    # Collect all clinical test files
    clinical_files = []
    for d in [FILES1_DIR, FILES2_DIR]:
        if os.path.isdir(d):
            for f in sorted(os.listdir(d)):
                if f.endswith(".xlsx"):
                    clinical_files.append(os.path.join(d, f))

    print(f"Found {len(clinical_files)} clinical test files + 1 TCP file\n")

    # Import clinical tests
    success = 0
    errors = []
    for filepath in clinical_files:
        filename = os.path.basename(filepath)
        print(f"Processing: {filename}")
        try:
            test_data = parse_clinical_excel(filepath)
            insert_test(conn, test_data)
            success += 1
        except Exception as e:
            print(f"  ERROR: {e}")
            errors.append((filename, str(e)))
            conn.rollback()

    # Import TCP test
    if os.path.exists(TCP_FILE):
        print(f"\nProcessing: TCP_test_final_solo.xlsx")
        try:
            test_data = parse_tcp_excel(TCP_FILE)
            insert_test(conn, test_data, code_override="TCP-FINAL")
            success += 1
        except Exception as e:
            print(f"  ERROR: {e}")
            errors.append(("TCP_test_final_solo.xlsx", str(e)))
            conn.rollback()

    print(f"\n{'='*60}")
    print(f"Results: {success} imported, {len(errors)} errors")
    if errors:
        print("\nErrors:")
        for fname, err in errors:
            print(f"  - {fname}: {err}")

    # Verify
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM tests")
    total_tests = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM questions")
    total_questions = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM answers")
    total_answers = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM subfactors")
    total_subfactors = cur.fetchone()[0]
    print(f"\nDB totals: {total_tests} tests, {total_questions} questions, {total_answers} answers, {total_subfactors} subfactors")
    cur.close()

    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
